from __future__ import annotations

import json
import warnings
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import urlparse

try:
    warnings.filterwarnings("ignore", category=DeprecationWarning)
    import cgi
    import openpyxl
except Exception:  # pragma: no cover - local runtime fallback
    import cgi
    openpyxl = None


ROOT = Path(__file__).resolve().parent
PORTAL = ROOT / "INDANE_SALES_MONITORING_PORTAL.html"
SEED = ROOT / "data" / "portal-seed.json"


def load_seed() -> dict:
    return json.loads(SEED.read_text(encoding="utf-8"))


def save_seed(seed: dict) -> None:
    SEED.write_text(json.dumps(seed, ensure_ascii=False, indent=2), encoding="utf-8")


def parse_distributor_workbook(file_bytes: bytes, filename: str) -> list[dict]:
    if filename.lower().endswith(".csv"):
        import csv
        import io

        reader = csv.DictReader(io.StringIO(file_bytes.decode("utf-8-sig", errors="ignore")))
        return [normalize_distributor(row) for row in reader if normalize_distributor(row).get("sap")]

    if openpyxl is None:
        raise RuntimeError("openpyxl is not available for .xlsx upload on this machine")

    import io

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(rows)]
    output: list[dict] = []
    for values in rows:
        row = dict(zip(headers, values))
        normalized = normalize_distributor(row)
        if normalized.get("sap"):
            output.append(normalized)
    return output


def parse_spd_workbook(file_bytes: bytes, filename: str) -> list[dict]:
    if filename.lower().endswith(".csv"):
        import csv
        import io

        reader = csv.DictReader(io.StringIO(file_bytes.decode("utf-8-sig", errors="ignore")))
        return [normalize_spd(row) for row in reader if normalize_spd(row).get("sap")]

    if openpyxl is None:
        raise RuntimeError("openpyxl is not available for .xlsx upload on this machine")

    import io

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(rows)]
    output: list[dict] = []
    for values in rows:
        row = dict(zip(headers, values))
        normalized = normalize_spd(row)
        if normalized.get("sap"):
            output.append(normalized)
    return output


def parse_invoice_workbook(file_bytes: bytes, filename: str) -> list[dict]:
    text_probe = ""
    for encoding in ("utf-16", "utf-8-sig", "utf-8"):
        try:
            text_probe = file_bytes[:4096].decode(encoding)
            break
        except UnicodeError:
            continue

    if "Invoice No" in text_probe and "\t" in text_probe:
        import csv
        import io

        text = file_bytes.decode("utf-16" if file_bytes.startswith(b"\xff\xfe") else "utf-8-sig", errors="ignore")
        lines = [line for line in text.splitlines() if line.strip()]
        header_index = next((idx for idx, line in enumerate(lines) if "Invoice No" in line and "Customer" in line), 0)
        rows = list(csv.DictReader(io.StringIO("\n".join(lines[header_index:])), delimiter="\t"))
    elif filename.lower().endswith(".csv"):
        import csv
        import io

        rows = list(csv.DictReader(io.StringIO(file_bytes.decode("utf-8-sig", errors="ignore"))))
    else:
        if openpyxl is None:
            raise RuntimeError("openpyxl is not available for .xlsx upload on this machine")
        import io

        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        raw_rows = ws.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(raw_rows)]
        rows = [dict(zip(headers, values)) for values in raw_rows]

    invoices: dict[str, dict] = {}
    for row in rows:
        joined = " ".join(str(value or "") for value in row.values())
        if "CALC" in joined.upper() or "CANCEL" in joined.upper():
            continue
        invoice_no = pick(row, "Billing Document", "Invoice No.", "Invoice No", "Invoice", "Document No") or f"INV-{len(invoices)+1}"
        sap_code = pick(row, "Customer", "SAP Code", "Distributor SAP Code")
        distributor_name = pick(row, "Ship-To Party", "Sold-To Party", "Distributor Name", "Name")
        if not distributor_name:
            distributor_name = pick(row, "Customer Name")
        material_text = pick(row, "Material", "Material Description", "Item Description")
        if material_text and "14.2" not in material_text:
            continue
        plant = pick(row, "Plant", "Shipping Point/Receiving Pt", "Supply Plant")
        if not plant:
            plant = pick(row, "Plant Name")
        date_value = pick(row, "Date", "Billing Date", "Invoice Date")
        vendor_code = pick(row, "Vendor", "Vendor Code", "Transporter")
        vendor_name = pick(row, "Vendor Name", "Transporter Name")
        if not distributor_name:
            continue
        key = f"{sap_code or distributor_name}|{invoice_no}"
        invoices[key] = {
            "sap": sap_code,
            "name": distributor_name,
            "vendor": vendor_code,
            "vendor_name": vendor_name,
            "invoice_no": invoice_no,
            "plant": plant,
            "date": date_value,
            "loads": 1,
            "source": "YVLPAUTO_Invoicing",
        }
    return list(invoices.values())


def pick(row: dict, *names: str) -> str:
    lowered = {str(k).strip().lower(): v for k, v in row.items()}
    for name in names:
        value = lowered.get(name.lower())
        if value is not None:
            return str(value).strip()
    compact = {str(k).strip().lower().replace(".", "").replace(" ", ""): v for k, v in row.items()}
    for name in names:
        value = compact.get(name.lower().replace(".", "").replace(" ", ""))
        if value is not None:
            return str(value).strip()
    return ""


def normalize_distributor(row: dict) -> dict:
    plant = pick(row, "Supply Plant", "Default Supply Plant", "Plant")
    ido = pick(row, "Sales Office", "IDO", "Sales Off.")
    if not ido:
        ido = "Dehradun IDO" if plant in {"Haridwar BP", "Kashipur BP"} else "Noida IDO"
    return {
        "sap": pick(row, "SAP Code", "SAP", "Distributor SAP Code"),
        "name": pick(row, "Distributor name", "Name of distributorship", "Distributor Name", "Name"),
        "email": pick(row, "Email", "Email ID", "Mail ID"),
        "address": pick(row, "Address", "Physical Address"),
        "district": pick(row, "District"),
        "lsa": pick(row, "LSA", "LPG Sales Area", "Sales Group"),
        "ido": ido,
        "plant": plant,
        "urban": pick(row, "Urban/Rural", "Urban Rural"),
        "loads": int(float(pick(row, "Final Planning", "SPD Loads", "Loads") or 0)),
    }


def normalize_spd(row: dict) -> dict:
    loads_raw = pick(row, "SPD Loads", "Final Planning", "Loads", "No. of Loads", "Target Loads")
    cyl_raw = pick(row, "SPD Cylinders", "Target Cylinders", "Cylinders")
    loads = int(float(loads_raw or 0))
    if not loads and cyl_raw:
        loads = int((float(cyl_raw) + 359) // 360)
    return {
        "date": pick(row, "Date", "Planning Date", "SPD Date"),
        "sap": pick(row, "SAP Code", "SAP", "Distributor SAP Code"),
        "loads": loads,
        "cyl": loads * 360,
        "priority": pick(row, "Priority", "Priority Level") or "Normal",
        "backlog": int(float(pick(row, "Backlog", "Backlog Cylinders", "Backlog Qty") or 0)),
        "indent": pick(row, "Indent No", "Indent", "SAP Indent No"),
        "reason": pick(row, "Reason", "Override Reason", "Remarks"),
    }


class PortalHandler(SimpleHTTPRequestHandler):
    def translate_path(self, path: str) -> str:
        requested = urlparse(path).path
        if requested == "/":
            return str(PORTAL)
        return str(ROOT / requested.lstrip("/"))

    def end_headers(self) -> None:
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Cache-Control", "no-store")
        super().end_headers()

    def do_GET(self) -> None:
        if self.path == "/api/seed":
            self.send_json(load_seed())
            return
        super().do_GET()

    def do_POST(self) -> None:
        if self.path == "/api/upload/distributor-master":
            self.handle_distributor_upload()
            return
        if self.path == "/api/upload/spd":
            self.handle_spd_upload()
            return
        if self.path == "/api/upload/invoices":
            self.handle_invoice_upload()
            return
        self.send_error(404, "Unknown endpoint")

    def handle_distributor_upload(self) -> None:
        form = cgi.FieldStorage(fp=self.rfile, headers=self.headers, environ={"REQUEST_METHOD": "POST"})
        item = form["file"] if "file" in form else None
        if item is None or not getattr(item, "filename", ""):
            self.send_json({"status": "error", "message": "No file uploaded"}, status=400)
            return
        try:
            uploaded = parse_distributor_workbook(item.file.read(), item.filename)
            seed = load_seed()
            current = {row["sap"]: row for row in seed.get("distributors", []) if row.get("sap")}
            for row in uploaded:
                current[row["sap"]] = {**current.get(row["sap"], {}), **row}
            seed["distributors"] = list(current.values())
            seed["dist_count"] = len(seed["distributors"])
            save_seed(seed)
            self.send_json({"status": "uploaded", "rows": len(uploaded), "total": len(seed["distributors"])})
        except Exception as exc:
            self.send_json({"status": "error", "message": str(exc)}, status=500)

    def handle_invoice_upload(self) -> None:
        form = cgi.FieldStorage(fp=self.rfile, headers=self.headers, environ={"REQUEST_METHOD": "POST"})
        item = form["file"] if "file" in form else None
        if item is None or not getattr(item, "filename", ""):
            self.send_json({"status": "error", "message": "No file uploaded"}, status=400)
            return
        try:
            invoices = parse_invoice_workbook(item.file.read(), item.filename)
            self.send_json({"status": "uploaded", "rows": len(invoices), "invoices": invoices})
        except Exception as exc:
            self.send_json({"status": "error", "message": str(exc)}, status=500)

    def handle_spd_upload(self) -> None:
        form = cgi.FieldStorage(fp=self.rfile, headers=self.headers, environ={"REQUEST_METHOD": "POST"})
        item = form["file"] if "file" in form else None
        if item is None or not getattr(item, "filename", ""):
            self.send_json({"status": "error", "message": "No file uploaded"}, status=400)
            return
        try:
            uploaded = parse_spd_workbook(item.file.read(), item.filename)
            seed = load_seed()
            seed["uploaded_spd"] = uploaded
            save_seed(seed)
            self.send_json({"status": "uploaded", "rows": len(uploaded)})
        except Exception as exc:
            self.send_json({"status": "error", "message": str(exc)}, status=500)

    def send_json(self, payload: dict, status: int = 200) -> None:
        data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)


if __name__ == "__main__":
    server = ThreadingHTTPServer(("0.0.0.0", 8095), PortalHandler)
    print("INDANE SALES MONITORING running at http://127.0.0.1:8095")
    print("Keep this window open. Press Ctrl+C to stop.")
    server.serve_forever()
